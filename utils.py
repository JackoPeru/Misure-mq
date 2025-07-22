# utils.py

import openpyxl
import xlrd
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import json
import database
from datetime import datetime

def export_to_excel(quote_data, summary_data):
    """Esporta i dati del preventivo e il riepilogo in un file Excel."""
    pass # Implementazione futura

def export_to_pdf(quote_items, totals, filename):
    """Esporta il preventivo in un file PDF."""
    doc = SimpleDocTemplate(filename, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Preventivo Soglie", styles['h1']))
    story.append(Spacer(1, 12))

    # Dati tabella
    data = [["N.", "L (cm)", "W (cm)", "Materiale", "Spess. (cm)", "m²", "€/m²", "Importo (€)"]]
    for item in quote_items:
        data.append(item)
    
    # Creazione tabella
    table = Table(data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ])
    table.setStyle(style)
    story.append(table)
    story.append(Spacer(1, 24))

    # Riepilogo
    story.append(Paragraph(f"<b>Totale m²:</b> {totals['mq']}", styles['Normal']))
    story.append(Paragraph(f"<b>Totale Importo (€):</b> {totals['eur']}", styles['Normal']))

    doc.build(story)
    messagebox.showinfo("Esportazione PDF", f"Preventivo esportato con successo in {filename}", parent=None) # Or pass parent if available

def save_quote_to_json(quote_items, totals, filename):
    """Salva il preventivo corrente in un file JSON."""
    data_to_save = {
        "items": quote_items,
        "totals": totals
    }
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data_to_save, f, indent=4, ensure_ascii=False)
        messagebox.showinfo("Salvataggio JSON", f"Preventivo salvato con successo in {filename}", parent=None)
    except IOError as e:
        messagebox.showerror("Errore Salvataggio", f"Impossibile salvare il file: {e}", parent=None)

def load_quote_from_json(filename):
    """Carica un preventivo da un file JSON."""
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data.get("items", []), data.get("totals", {})
    except FileNotFoundError:
        messagebox.showerror("Errore Apertura", "File non trovato.", parent=None)
        return None, None
    except json.JSONDecodeError:
        messagebox.showerror("Errore Apertura", "Formato file JSON non valido.", parent=None)
        return None, None
    except IOError as e:
        messagebox.showerror("Errore Apertura", f"Impossibile leggere il file: {e}", parent=None)
        return None, None

def import_materials_from_excel(filepath, progress_callback=None):
    """Importa materiali da un file Excel multi-sheet nel database SQLite."""
    imported_count = 0
    skipped_count = 0
    processed_sheets = 0

    try:
        if filepath.endswith('.xls'):
            workbook = xlrd.open_workbook(filepath)
            sheet_names = workbook.sheet_names()
        elif filepath.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(filepath, data_only=True) # data_only=True to get values not formulas
            sheet_names = workbook.sheetnames
        else:
            messagebox.showerror("Errore Importazione", "Formato file non supportato. Utilizzare .xls o .xlsx", parent=None)
            return False

        total_sheets = len(sheet_names)
        if total_sheets == 0:
            messagebox.showerror("Errore Importazione", "Il file Excel non contiene fogli.", parent=None)
            return False

        for sheet_index, supplier_name in enumerate(sheet_names):
            if progress_callback:
                # Update progress based on sheets processed, and then rows within a sheet
                progress_callback((sheet_index / total_sheets) * 100) # Initial progress for starting a new sheet
            
            if filepath.endswith('.xls'):
                sheet = workbook.sheet_by_name(supplier_name)
                # Find the header row (e.g., containing 'nome' or 'Nome Materiale')
                header_row_idx = -1
                for r_idx in range(sheet.nrows):
                    row_values = [str(sheet.cell_value(r_idx, c_idx)).strip().lower() for c_idx in range(sheet.ncols)]
                    if "nome" in row_values or "nome materiale" in row_values:
                        header_row_idx = r_idx
                        break
                if header_row_idx == -1:
                    print(f"Foglio '{supplier_name}': Intestazione non trovata. Salto.")
                    skipped_count += sheet.nrows # Approximate skipped rows
                    continue
                
                header = [str(sheet.cell_value(header_row_idx, col)).strip() for col in range(sheet.ncols)]
                data_rows = (sheet.row_values(row_idx) for row_idx in range(header_row_idx + 1, sheet.nrows))
                current_sheet_total_rows = sheet.nrows - (header_row_idx + 1)
            
            elif filepath.endswith('.xlsx'):
                sheet = workbook[supplier_name]
                header_row_idx = -1
                for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    row_values = [str(cell).strip().lower() if cell is not None else "" for cell in row]
                    if "nome" in row_values or "nome materiale" in row_values:
                        header_row_idx = r_idx
                        break
                if header_row_idx == -1:
                    print(f"Foglio '{supplier_name}': Intestazione non trovata. Salto.")
                    skipped_count += sheet.max_row # Approximate skipped rows
                    continue

                # Get header row values more robustly
                header_values_generator = sheet.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 1, values_only=True)
                header = []
                for row_tuple in header_values_generator: # Should be only one row_tuple
                    header = [str(cell).strip() if cell is not None else "" for cell in row_tuple]
                    break # exit after processing the header row
                print(f"Debug: Foglio '{supplier_name}', Header identificato: {header}") # Debug print

                data_rows = sheet.iter_rows(min_row=header_row_idx + 2, values_only=True)
                current_sheet_total_rows = sheet.max_row - (header_row_idx + 1)

            if not header or not any(h.lower() == "nome materiale" or h.lower() == "nome" for h in header):
                print(f"Foglio '{supplier_name}': Colonna 'Nome Materiale' non trovata nell'intestazione. Salto.")
                skipped_count += current_sheet_total_rows
                continue
            
            try:
                name_col_idx = next(i for i, h in enumerate(header) if h.lower() == "nome materiale" or h.lower() == "nome")
            except StopIteration:
                print(f"Foglio '{supplier_name}': Indice colonna 'Nome Materiale' non trovato. Salto.")
                skipped_count += current_sheet_total_rows
                continue

            # Identify thickness columns (e.g., CM2, CM3, CM40)
            thickness_price_cols = {}
            for col_idx, col_name in enumerate(header):
                if col_name.upper().startswith("CM") and col_idx != name_col_idx:
                    try:
                        # Extract numeric part of thickness from CM<number>
                        thickness_val = float(col_name[2:].replace(',', '.'))
                        thickness_price_cols[thickness_val] = col_idx
                    except ValueError:
                        print(f"Foglio '{supplier_name}': Formato spessore non valido in colonna '{col_name}'. Salto colonna.")
            
            if not thickness_price_cols:
                print(f"Foglio '{supplier_name}': Nessuna colonna spessore (CMx) trovata o valida. Salto foglio.")
                skipped_count += current_sheet_total_rows
                continue

            sheet_imported_count = 0
            sheet_skipped_count = 0
            for i, row_data in enumerate(data_rows):
                if progress_callback:
                    # Update progress for rows within the current sheet
                    sheet_progress = ((i + 1) / current_sheet_total_rows) * (100 / total_sheets) # Progress within this sheet's share
                    base_progress = (sheet_index / total_sheets) * 100
                    progress_callback(base_progress + sheet_progress)

                material_name = str(row_data[name_col_idx]).strip() if row_data[name_col_idx] else None

                if not material_name:
                    sheet_skipped_count += len(thickness_price_cols) # Skipping all potential entries for this row
                    continue

                for thickness, price_col_idx in thickness_price_cols.items():
                    if price_col_idx >= len(row_data) or row_data[price_col_idx] is None or str(row_data[price_col_idx]).strip() == "":
                        # No price for this thickness, skip this specific material-thickness combination
                        sheet_skipped_count += 1
                        continue
                    
                    price_str = str(row_data[price_col_idx]).replace(',', '.').replace('€', '').strip()
                    
                    try:
                        price = float(price_str)
                    except ValueError:
                        print(f"Foglio '{supplier_name}', Materiale '{material_name}', Spessore '{thickness}': Prezzo non valido '{price_str}'. Salto.")
                        sheet_skipped_count += 1
                        continue
                    
                    # Description is not part of this specific Excel structure, so pass empty or None
                    description = ""

                    if database.add_material(material_name, price, thickness, description, supplier_name):
                        sheet_imported_count += 1
                    else:
                        sheet_skipped_count += 1
            
            imported_count += sheet_imported_count
            skipped_count += sheet_skipped_count
            processed_sheets += 1
            print(f"Foglio '{supplier_name}': Importati {sheet_imported_count}, Saltati {sheet_skipped_count}")

        if progress_callback:
            progress_callback(100) # Ensure progress reaches 100%

        messagebox.showinfo("Importazione Completata", 
                              f"{imported_count} materiali importati da {processed_sheets} fogli.\n"
                              f"{skipped_count} voci saltate (dati mancanti/errati o duplicati).", parent=None)
        return True

    except FileNotFoundError:
        messagebox.showerror("Errore Importazione", "File Excel non trovato.", parent=None)
        return False
    except xlrd.XLRDError as e:
        messagebox.showerror("Errore Importazione XLS", f"Errore durante la lettura del file .xls: {e}", parent=None)
        return False
    except openpyxl.utils.exceptions.InvalidFileException as e:
        messagebox.showerror("Errore Importazione XLSX", f"Errore durante la lettura del file .xlsx: {e}", parent=None)
        return False
    except Exception as e:
        import traceback
        traceback.print_exc() # For debugging
        messagebox.showerror("Errore Importazione", f"Si è verificato un errore imprevisto: {e}", parent=None)
        return False

# --- Helper for progress bar during import ---
class ProgressWindow(tk.Toplevel):
    def __init__(self, parent, title="Importazione..."):
        super().__init__(parent)
        self.title(title)
        self.geometry("300x100")
        self.resizable(False, False)

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self, variable=self.progress_var, maximum=100, length=280)
        self.progress_bar.pack(pady=20, padx=10)

        self.label_var = tk.StringVar(value="Inizio importazione...")
        self.label = ttk.Label(self, textvariable=self.label_var)
        self.label.pack()

        self.transient(parent)
        self.grab_set()
        self.update_idletasks() # Ensure window is drawn

    def update_progress(self, value):
        self.progress_var.set(value)
        self.label_var.set(f"Importazione: {value:.0f}%")
        if value >= 100:
            self.label_var.set("Completato!")
            self.after(1000, self.destroy) # Close after 1 sec
        self.update_idletasks()


def export_quote_to_pdf(filename, quote_data, totals, edge_details_for_pdf, client_name, quote_date, quote_number):
    doc = SimpleDocTemplate(filename, pagesize=letter,
                          rightMargin=30, leftMargin=30,
                          topMargin=30, bottomMargin=18)
    elements = []
    styles = getSampleStyleSheet()
    styleN = styles['Normal']
    styleH = styles['Heading2']
    styleB = styles['BodyText']
    styleI = styles['Italic']

    # Titolo e Info Preventivo
    elements.append(Paragraph(f"Preventivo N: {quote_number}", styleH))
    elements.append(Paragraph(f"Data: {quote_date}", styleN))
    elements.append(Paragraph(f"Cliente: {client_name if client_name else 'Non specificato'}", styleN))
    elements.append(Spacer(1, 0.5*cm))

    # Dati Tabella
    header = ["N.", "L (cm)", "W (cm)", "Materiale", "Spess. (cm)", "m²", "€/m²", "Lastra (€)", "Bordi (€)", "Tot. Riga (€)"]
    data_for_table = [header]

    for i, row_values in enumerate(quote_data):
        # Formatta i valori della riga principale
        formatted_row = [str(val) for val in row_values[:10]] # Prendi solo i primi 10 valori (escludi l'ID riga)
        data_for_table.append(formatted_row)

        # Se ci sono dettagli dei bordi per questa riga, aggiungili come sotto-riga
        if edge_details_for_pdf:
            # L'ID riga dovrebbe essere l'ultimo elemento di row_values
            row_id = row_values[-1] if len(row_values) > 10 else None
            if row_id and row_id in edge_details_for_pdf:
                edge_detail = edge_details_for_pdf[row_id]
                if edge_detail and edge_detail.get('total_edge_cost', 0) > 0:
                    details_str_parts = []
                    for side_key in ['front', 'back', 'left', 'right']:
                        side_data = edge_detail.get(side_key, {})
                        if isinstance(side_data, dict) and side_data.get('active'):
                            side_name = {
                                'front': 'Fronte',
                                'back': 'Retro',
                                'left': 'Sinistra',
                                'right': 'Destra'
                            }.get(side_key, side_key.capitalize())
                            details_str_parts.append(
                                f"{side_name}: {side_data['type']} ({side_data['length_cm']:.1f}cm, {side_data['cost']:.2f}€)"
                            )
                    
                    if details_str_parts:
                        # Crea una riga di dettaglio con il testo dei bordi nella colonna del materiale
                        # e spazi vuoti nelle altre colonne
                        detail_row = [''] * len(header)
                        detail_row[3] = f"    Bordi: {'; '.join(details_str_parts)}" # Indenta e usa la colonna 'Materiale'
                        data_for_table.append(detail_row)

    # Crea e stila la tabella
    table = Table(data_for_table, colWidths=[0.8*cm, 1.8*cm, 1.8*cm, 4*cm, 2*cm, 1.8*cm, 1.8*cm, 2*cm, 2*cm, 2.2*cm])
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'), # Numeri centrati
        ('ALIGN', (1, 1), (2, -1), 'RIGHT'),  # Dimensioni allineate a destra
        ('ALIGN', (5, 1), (-1, -1), 'RIGHT'), # Valori numerici allineati a destra
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        # Stile per le righe di dettaglio dei bordi
        ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Oblique'), # Corsivo per i dettagli dei bordi
        ('FONTSIZE', (3, 1), (3, -1), 8),                    # Testo più piccolo per i dettagli
        ('TEXTCOLOR', (3, 1), (3, -1), colors.dark_grey),    # Colore più chiaro per i dettagli
    ])
    table.setStyle(table_style)
    elements.append(table)
    elements.append(Spacer(1, 0.5*cm))

    # Riepilogo Totali
    summary_data = [
        ["Totale m²:", totals.get('total_mq', '0.0000')],
        ["Costo Lastre (€):", totals.get('total_slabs_eur', '0.00')],
        ["Costo Bordi (€):", totals.get('total_edges_eur', '0.00')],
        [Paragraph("<b>Totale Complessivo (€):</b>", styleN),
         Paragraph(f"<b>{totals.get('total_eur', '0.00')}</b>", styleN)]
    ]
    summary_table = Table(summary_data, colWidths=[5*cm, 3*cm])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, -1), (-1, -1), 1, colors.black),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(summary_table)

    # Genera il PDF
    doc.build(elements)
    messagebox.showinfo("Esportazione PDF", f"Preventivo esportato con successo in {filename}")


def export_materials_and_edges_to_json(filename):
    """Esporta tutti i materiali e tipi di bordo in un file JSON per backup/trasferimento."""
    try:
        # Recupera tutti i materiali
        materials = database.get_all_materials()
        materials_data = []
        for material in materials:
            materials_data.append({
                'name': material['name'],
                'thickness': material['thickness'],
                'price_per_sqm': material['price_per_sqm'],
                'description': material['description'],
                'supplier': material['supplier']
            })
        
        # Recupera tutti i tipi di bordo
        edges = database.get_all_edge_types()
        edges_data = []
        for edge in edges:
            edges_data.append({
                'material_name': edge['material_name'],
                'thickness': edge['thickness'],
                'edge_type': edge['edge_type'],
                'price_per_lm': edge['price_per_lm']
            })
        
        # Crea il dizionario da esportare
        export_data = {
            'export_date': json.dumps(datetime.now().isoformat()),
            'materials': materials_data,
            'edges': edges_data
        }
        
        # Salva nel file JSON
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=4, ensure_ascii=False)
        
        messagebox.showinfo("Esportazione Completata", 
                          f"Esportati {len(materials_data)} materiali e {len(edges_data)} tipi di bordo in {filename}")
        return True
        
    except Exception as e:
        messagebox.showerror("Errore Esportazione", f"Errore durante l'esportazione: {e}")
        return False


def import_materials_and_edges_from_json(filename, overwrite_existing=False):
    """Importa materiali e tipi di bordo da un file JSON di backup."""
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            import_data = json.load(f)
        
        if 'materials' not in import_data and 'edges' not in import_data:
            messagebox.showerror("Errore Importazione", "File JSON non valido: mancano le sezioni 'materials' o 'edges'")
            return False
        
        imported_materials = 0
        skipped_materials = 0
        imported_edges = 0
        skipped_edges = 0
        
        # Importa materiali
        if 'materials' in import_data:
            for material_data in import_data['materials']:
                try:
                    result = database.add_material(
                        name=material_data['name'],
                        price_per_sqm=material_data['price_per_sqm'],
                        thickness=material_data.get('thickness'),
                        description=material_data.get('description', ''),
                        supplier=material_data.get('supplier')
                    )
                    if result:
                        imported_materials += 1
                    else:
                        skipped_materials += 1
                except Exception as e:
                    print(f"Errore importazione materiale {material_data.get('name', 'sconosciuto')}: {e}")
                    skipped_materials += 1
        
        # Importa tipi di bordo
        if 'edges' in import_data:
            for edge_data in import_data['edges']:
                try:
                    result = database.add_edge_type(
                        edge_type=edge_data['edge_type'],
                        price_per_lm=edge_data['price_per_lm'],
                        material_name=edge_data.get('material_name'),
                        thickness=edge_data.get('thickness')
                    )
                    if result:
                        imported_edges += 1
                    else:
                        skipped_edges += 1
                except Exception as e:
                    print(f"Errore importazione bordo {edge_data.get('edge_type', 'sconosciuto')}: {e}")
                    skipped_edges += 1
        
        messagebox.showinfo("Importazione Completata", 
                          f"Importati: {imported_materials} materiali, {imported_edges} tipi di bordo\n"
                          f"Saltati (duplicati/errori): {skipped_materials} materiali, {skipped_edges} tipi di bordo")
        return True
        
    except FileNotFoundError:
        messagebox.showerror("Errore Importazione", "File non trovato")
        return False
    except json.JSONDecodeError:
        messagebox.showerror("Errore Importazione", "File JSON non valido")
        return False
    except Exception as e:
        messagebox.showerror("Errore Importazione", f"Errore durante l'importazione: {e}")
        return False